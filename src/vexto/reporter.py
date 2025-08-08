# src/vexto/reporter.py

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import ast # Nødvendig for at læse link-data sikkert

def create_excel_report(df_nested: pd.DataFrame):
    """
    Genererer en komplet og poleret Excel-rapport (v2.2) med separat fane
    for ødelagte links og forbedret læsbarhed.
    """
    # --- TRIN 0: FORBEREDELSE & FILNAVN ---
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    path = f"output/Vexto_Rapport_{timestamp}.xlsx"
    Path("output").mkdir(exist_ok=True)

    # --- TRIN 1: DATA FORBEREDELSE (Udpakning af nestede kolonner) ---
    print("Forbereder data og pakker nestede kolonner ud...")
    
    nested_columns = [
        'basic_seo', 'technical_seo', 'performance', 'authority', 
        'security', 'privacy', 'conversion', 'social_and_reputation'
    ]
    
    df_flat = pd.DataFrame(df_nested['url'])

    for col_name in nested_columns:
        if col_name in df_nested.columns and df_nested[col_name].notna().any():
            valid_rows = df_nested[df_nested[col_name].apply(isinstance, args=(dict,))]
            if not valid_rows.empty:
                normalized_data = pd.json_normalize(valid_rows[col_name]).add_prefix(f'{col_name}_')
                normalized_data.index = valid_rows.index
                df_flat = df_flat.join(normalized_data)
    
    print("Data er færdigbehandlet.")

    # --- TRIN 2: OPRET FANER ---
    with pd.ExcelWriter(path, engine="openpyxl") as xls:
        # --- ÆNDRING 1: Fjern "Døde Links Liste" fra Technical SEO fanen ---
        sheet_definitions = {
            "Overview": {
                "URL": "url",
                "Titel": "basic_seo_title_text",
                "Performance Score": "performance_performance_score",
                "Authority Rank": "authority_global_rank",
                "Døde Links": "technical_seo_broken_links_count",
                "Sikkerheds-tjek": "security_hsts_enabled"
            },
            "Technical SEO": {
                "URL": "url",
                "HTTPS": "technical_seo_is_https",
                "Robots.txt Fundet": "technical_seo_robots_txt_found",
                "Sitemap Fundet": "technical_seo_sitemap_xml_found",
                "Døde Links Antal": "technical_seo_broken_links_count",
                # "Døde Links Liste" er fjernet herfra for læsbarhed
            },
            "Performance": {
                "URL": "url",
                "PSI Score": "performance_performance_score",
                "LCP (ms)": "performance_lcp_ms",
                "Total JS (KB)": "performance_total_js_size_kb",
                "Gns. Billedstr. (KB)": "basic_seo_avg_image_size_kb"
            },
            "Conversion & Privacy": {
                "URL": "url",
                "GA4 Fundet": "conversion_has_ga4",
                "Meta Pixel Fundet": "conversion_has_meta_pixel",
                "Cookie Banner": "privacy_cookie_banner_detected",
                "Fundne E-mails": "conversion_emails_found",
                "Fundne Tlf.nr.": "conversion_phone_numbers_found",
                "Trust Signals": "conversion_trust_signals_found"
            }
        }

        # Opret hver fane fra definitionerne
        for sheet_name, cols_map in sheet_definitions.items():
            existing_cols = {display_name: actual_name for display_name, actual_name in cols_map.items() if actual_name in df_flat.columns}
            if existing_cols:
                sheet_df = df_flat[list(existing_cols.values())].copy()
                sheet_df.columns = list(existing_cols.keys())
                sheet_df.to_excel(xls, sheet_name=sheet_name, index=False)
        
        # --- ÆNDRING 2: Opret en ny, separat fane for ødelagte links ---
        broken_links_data = []
        if 'technical_seo_broken_links_list' in df_flat.columns:
            for _, row in df_flat.iterrows():
                if pd.notna(row['technical_seo_broken_links_list']):
                    try:
                        # Konverter strengen sikkert til et dictionary
                        links_dict = ast.literal_eval(row['technical_seo_broken_links_list'])
                        if isinstance(links_dict, dict):
                            for link, status in links_dict.items():
                                broken_links_data.append({
                                    'Analyseret URL': row['url'],
                                    'Ødelagt Link': link,
                                    'Statuskode': status
                                })
                    except (ValueError, SyntaxError):
                        continue # Ignorer, hvis data ikke kan læses

        if broken_links_data:
            broken_links_df = pd.DataFrame(broken_links_data)
            broken_links_df.to_excel(xls, sheet_name="Broken Links Detaljer", index=False)
        
        # Gem en fane med alle udpakkede data for fuld gennemsigtighed
        df_flat.to_excel(xls, "Alle Data", index=False)

        # --- TRIN 3: VISUEL POLISH (Uændret) ---
        wb = xls.book
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="003366")
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        green_fill = PatternFill("solid", fgColor="C6EFCE")

        for ws in wb.worksheets:
            if ws.max_row == 0: continue
            
            # Style header
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(vertical="center", horizontal="center")

            # Autofit kolonnebredder
            for col_idx, column in enumerate(ws.columns, 1):
                max_length = max(len(str(cell.value or "")) for cell in column)
                max_length = max(max_length, len(ws.cell(row=1, column=col_idx).value or ""))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

        # Specifik conditional formatting
        if "Overview" in wb.sheetnames:
            ws = wb["Overview"]
            ws.conditional_formatting.add(f"C2:C{ws.max_row}", ColorScaleRule(start_type='num', start_value=0, start_color='FFC7CE', mid_type='num', mid_value=70, mid_color='FFEB9C', end_type='num', end_value=100, end_color='C6EFCE'))
            ws.conditional_formatting.add(f"E2:E{ws.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=red_fill))
            ws.conditional_formatting.add(f"F2:F{ws.max_row}", CellIsRule(operator="equal", formula=["TRUE"], fill=green_fill))

    print(f"FÆRDIG: Rapport 2.2 er genereret og gemt som '{path}'")