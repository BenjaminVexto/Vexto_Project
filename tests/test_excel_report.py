"""
test_excel_report.py
--------------------
Et script til at teste, hvor komplekst det er at generere en formateret
Excel-rapport ved hjælp af openpyxl.

Dette er en "spike" for at afklare indsatsen for den endelige rapportering.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Opsætning af stier ---
PROJECT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = PROJECT_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
OUTPUT_FILE = OUTPUT_DIR / "test_rapport.xlsx"

# --- Dummy-data, der simulerer et scoringsresultat ---
dummy_data = [
    {'kriterie': 'HTTPS er aktiveret', 'max_point': 10, 'opnået_point': 10, 'kommentar': 'Perfekt. Siden er sikker.'},
    {'kriterie': 'Sidens Title Tag', 'max_point': 5, 'opnået_point': 2, 'kommentar': 'Title er for lang (75 tegn). Bør være under 65.'},
    {'kriterie': 'Meta Description', 'max_point': 5, 'opnået_point': 0, 'kommentar': 'Siden mangler en meta description.'},
    {'kriterie': 'Core Web Vitals (LCP)', 'max_point': 10, 'opnået_point': 8, 'kommentar': 'LCP er 2.1s. Godkendt, men kan forbedres.'},
    {'kriterie': 'Billedoptimering', 'max_point': 5, 'opnået_point': 5, 'kommentar': 'Alle billeder bruger moderne formater (WebP).'},
]

# --- Definition af styles (farver, skrifttyper, kanter) ---
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

good_score_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
bad_score_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def create_formatted_report():
    """
    Opretter en formateret Excel-rapport fra bunden.
    """
    print(f"Opretter ny Excel-projektmappe...")
    # Opret en ny workbook og vælg det aktive ark
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scoringsrapport"

    # --- Trin 1: Skriv overskrifter og giv dem stil ---
    headers = ["Kriterie", "Maks Point", "Opnået Point", "Kommentar"]
    ws.append(headers)

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- Trin 2: Skriv data og anvend betinget formatering ---
    for data_row in dummy_data:
        # Konverter dict til en liste i korrekt rækkefølge
        row_values = [data_row['kriterie'], data_row['max_point'], data_row['opnået_point'], data_row['kommentar']]
        ws.append(row_values)
        
        # Anvend betinget formatering på "Opnået Point"-cellen
        current_row = ws.max_row
        opnået_cell = ws.cell(row=current_row, column=3)
        
        if data_row['opnået_point'] < data_row['max_point'] / 2:
            opnået_cell.fill = bad_score_fill # Rød farve for lav score
        else:
            opnået_cell.fill = good_score_fill # Grøn farve for god score

    # --- Trin 3: Anvend generel formatering på hele tabellen ---
    print("Anvender generel formatering (kanter, justering, kolonnebredde)...")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            # Juster tekst for de forskellige kolonner
            if cell.column in [2, 3]: # Maks og Opnået Point
                cell.alignment = center_align
            else: # Kriterie og Kommentar
                cell.alignment = left_align

    # --- Trin 4: Juster kolonnebredder ---
    ws.column_dimensions['A'].width = 40  # Kriterie
    ws.column_dimensions['B'].width = 15  # Maks Point
    ws.column_dimensions['C'].width = 15  # Opnået Point
    ws.column_dimensions['D'].width = 60  # Kommentar

    # --- Trin 5: Gem filen ---
    try:
        wb.save(OUTPUT_FILE)
        print(f"✓ Succes! Test-rapport gemt i: {OUTPUT_FILE}")
    except Exception as e:
        print(f"Fejl under gemning af fil: {e}")


if __name__ == "__main__":
    create_formatted_report()
